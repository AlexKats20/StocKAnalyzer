# streamlit_app.py

import streamlit as st
import yfinance as yf
import matplotlib.pyplot as plt
import openpyxl
import os
import numpy as np
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage
from matplotlib.gridspec import GridSpec
from matplotlib.patches import Rectangle
from mplfinance.original_flavor import candlestick_ohlc
import matplotlib.dates as mdates
import pandas as pd
import pandas_ta as ta
from fpdf import FPDF
from sklearn.linear_model import LinearRegression

# === Custom Style & Palette ===
tplt = plt.get_cmap('tab10')
PALETTE = tplt.colors
ACCENT = '#ff6f00'
plt.style.use('default')

# === Pattern Visual Highlight ===
def draw_pattern_visual(ax, df, idx, pattern_name):
    bar_width = 0.6
    if idx <= 1 or idx >= len(df):
        return
    involved = ([idx-2, idx-1, idx] if '3' in pattern_name or 'MORNINGSTAR' in pattern_name else [idx-1, idx])
    for i in involved:
        if 0 <= i < len(df):
            x0 = df['Date_Num'].iloc[i] - bar_width/2
            y0 = df['Low'].iloc[i]
            h = df['High'].iloc[i] - df['Low'].iloc[i]
            rect = Rectangle((x0, y0), bar_width, h, edgecolor=ACCENT, facecolor=ACCENT, alpha=0.4)
            ax.add_patch(rect)

# === Channel Detection ===
def detect_valid_channels(df, ax1, lookback=50, stride=5, min_slope=0.005):
    if len(df) < lookback:
        return
    last_channel_end = -1
    up_plotted = False
    down_plotted = False
    for start in range(0, len(df)-lookback, stride):
        if start < last_channel_end:
            continue
        end = start + lookback
        x = np.arange(lookback).reshape(-1,1)
        y_high = df['High'].iloc[start:end].values.reshape(-1,1)
        y_low = df['Low'].iloc[start:end].values.reshape(-1,1)
        close_prices = df['Close'].iloc[start:end].values

        reg_high = LinearRegression().fit(x, y_high)
        reg_low = LinearRegression().fit(x, y_low)
        slope_high, slope_low = reg_high.coef_[0][0], reg_low.coef_[0][0]

        upper_line = reg_high.predict(x).flatten()
        lower_line = reg_low.predict(x).flatten()
        channel_width = upper_line - lower_line
        width_var = np.std(channel_width)
        within_channel = np.mean((close_prices >= lower_line) & (close_prices <= upper_line))

        is_up = slope_high > min_slope and slope_low > min_slope
        is_down = slope_high < -min_slope and slope_low < -min_slope

        if (is_up or is_down) and width_var < 4.0 and within_channel > 0.4:
            color = PALETTE[3]
            hatch = '///'
            label = None
            if is_up and not up_plotted:
                label = 'Upward Channel'
                up_plotted = True
            if is_down and not down_plotted:
                label = 'Downward Channel'
                down_plotted = True

            date_nums = df['Date_Num'].iloc[start:end]
            ax1.plot(date_nums, upper_line, '--', lw=1.2, color=color, label=label)
            ax1.plot(date_nums, lower_line, '--', lw=1.2, color=color)
            ax1.fill_between(date_nums, lower_line, upper_line, color=color, alpha=0.2, hatch=hatch)
            last_channel_end = end

# === Stock Analysis ===
def analyze_stock(ticker, period, freq_str):
    freq_map = {'daily':'1d', 'weekly':'1wk', 'monthly':'1mo'}
    interval = freq_map.get(freq_str.lower(), '1d')
    actual_period = 'max' if str(period).lower() == 'max' else period

    print(f"⏳ Downloading data for {ticker} ({actual_period}, {freq_str})")
    df = yf.Ticker(ticker).history(period=actual_period, interval=interval,
                                   auto_adjust=False, prepost=True)
    df = df[['Open','High','Low','Close','Volume']].dropna()
    df['Date_Num'] = mdates.date2num(df.index.to_pydatetime())

    # Indicators
    df['MA20'] = df['Close'].rolling(20, min_periods=1).mean()
    df['MA50'] = df['Close'].rolling(50, min_periods=1).mean()
    df['MA100'] = df['Close'].rolling(100, min_periods=1).mean()

    delta = df['Close'].diff()
    gain = delta.where(delta>0,0)
    loss = -delta.where(delta<0,0)
    avg_gain = gain.rolling(14).mean()
    avg_loss = loss.rolling(14).mean()
    df['RSI'] = 100 - (100/(1 + avg_gain/avg_loss))

    ema12 = df['Close'].ewm(span=12).mean()
    ema26 = df['Close'].ewm(span=26).mean()
    df['MACD'] = ema12 - ema26
    df['Signal'] = df['MACD'].ewm(span=9).mean()

    # Pattern detection
    matches = []
    for name in dir(talib):
        if name.startswith('CDL'):
            result = getattr(talib, name)(df['Open'], df['High'], df['Low'], df['Close'])
            idxs = np.where(result != 0)[0]
            for idx in idxs:
                if (df.index[-1] - df.index[idx]).days <= 10:
                    matches.append({
                        'name': name.replace('CDL',''),
                        'index': idx,
                        'date': df.index[idx].date(),
                        'strength': int(result.iloc[idx])
                    })

    # Pattern fallback
    if matches:
        detected_pattern = matches[-1]['name']
        strength = f"{detected_pattern}={matches[-1]['strength']}"
    else:
        window = min(60, len(df))
        xw = np.arange(window).reshape(-1,1)
        yh = df['High'].iloc[-window:].values.reshape(-1,1)
        yl = df['Low'].iloc[-window:].values.reshape(-1,1)
        if window < 2:
            detected_pattern = 'None'; strength = ''
        else:
            rh = LinearRegression().fit(xw, yh)
            rl = LinearRegression().fit(xw, yl)
            sh, sl = rh.coef_[0][0], rl.coef_[0][0]
            if sh > 0.01 and sl > 0.01:
                detected_pattern = 'UpwardChannel'
            elif sh < -0.01 and sl < -0.01:
                detected_pattern = 'DownwardChannel'
            else:
                detected_pattern = 'None'
            strength = ''

    # Final classification
    rsi_val = df['RSI'].iloc[-1]
    macd_val = df['MACD'].iloc[-1]
    sig_val = df['Signal'].iloc[-1]
    classification = (
        'Bullish' if rsi_val<35 or (macd_val>sig_val and macd_val>-0.5)
        else 'Bearish' if rsi_val>65 or (macd_val<sig_val and macd_val<0.5)
        else 'Neutral'
    )

    # === Bullish markers ===
    bullish_indices = []

    for match in matches:
        pattern = match['name'].upper()
        idx = match['index']
        if 'HAMMER' in pattern or 'ENGULFING' in pattern or 'MORNINGSTAR' in pattern:
            if match['strength'] > 0:
                bullish_indices.append(idx)

    if rsi_val < 30:
        bullish_indices.append(len(df)-1)

    if df['MA20'].iloc[-1] > df['MA50'].iloc[-1] and df['MA20'].iloc[-2] < df['MA50'].iloc[-2]:
        bullish_indices.append(len(df)-1)

    # Plot window
    df_plot = df.tail(250 if interval=='1d' else 52 if interval=='1wk' else 24) if actual_period != 'max' else df

    chart_dir = 'charts'
    os.makedirs(chart_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    chart_path = f"{chart_dir}/{ticker}_{ts}.png"

    fig = plt.figure(figsize=(12,8))
    gs = GridSpec(4,1, height_ratios=[3,1,1,1])
    ax1 = fig.add_subplot(gs[0])
    ohlc = df_plot[['Date_Num','Open','High','Low','Close']].values
    candlestick_ohlc(ax1, ohlc, width=0.6, colorup=PALETTE[0], colordown=PALETTE[1])

    ax1.plot(df_plot['Date_Num'], df_plot['MA20'], color=PALETTE[0], lw=2, label='MA20')
    ax1.plot(df_plot['Date_Num'], df_plot['MA50'], color=PALETTE[1], lw=1.5, linestyle='--', label='MA50')
    ax1.plot(df_plot['Date_Num'], df_plot['MA100'], color=PALETTE[2], lw=1.5, linestyle=':', label='MA100')
    ax1.set_title(f"{ticker} | {freq_str.capitalize()} Chart")
    ax1.set_ylabel('Price')

    detect_valid_channels(df_plot, ax1)

    if detected_pattern and detected_pattern != 'None':
        raw_idx = matches[-1]['index'] if matches else len(df) - 1
        plot_start = len(df) - len(df_plot)
        if raw_idx >= plot_start:
            idx_plot = raw_idx - plot_start
            x0 = df_plot['Date_Num'].iloc[idx_plot]
            y0 = df_plot['High'].iloc[idx_plot]
            ax1.annotate(detected_pattern, xy=(x0, y0), xytext=(x0, y0 * 1.05),
                         bbox=dict(boxstyle='round,pad=0.3', fc='white', ec='black'),
                         arrowprops=dict(arrowstyle='->'))
            draw_pattern_visual(ax1, df_plot, idx_plot, detected_pattern)

    # === Plot all bullish markers ===
    plot_start = len(df) - len(df_plot)
    for idx in bullish_indices:
        if idx >= plot_start and idx < len(df):
            idx_plot = idx - plot_start
            x0 = df_plot['Date_Num'].iloc[idx_plot]
            y0 = df_plot['Low'].iloc[idx_plot] * 0.99
            ax1.scatter(x0, y0, color='green', marker='^', s=80, zorder=5)

    ax1.grid(True, linestyle=':', lw=0.5, alpha=0.4)
    ax1.legend(loc='upper left', bbox_to_anchor=(1.02,1))

    ax2 = fig.add_subplot(gs[1], sharex=ax1)
    ax2.bar(df_plot['Date_Num'], df_plot['Volume'], color='gray')

    ax3 = fig.add_subplot(gs[2], sharex=ax1)
    ax3.plot(df_plot['Date_Num'], df_plot['RSI'], color=PALETTE[4], label='RSI')
    ax3.axhline(70, color='red', linestyle='--')
    ax3.axhline(30, color='green', linestyle='--')
    ax3.legend(loc='upper left')

    ax4 = fig.add_subplot(gs[3], sharex=ax1)
    ax4.plot(df_plot['Date_Num'], df_plot['MACD'], color=PALETTE[5], label='MACD')
    ax4.plot(df_plot['Date_Num'], df_plot['Signal'], color=PALETTE[6], label='Signal')
    ax4.legend(loc='upper left')
    ax4.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))

    fig.autofmt_xdate(); plt.tight_layout()
    plt.savefig(chart_path, dpi=100); plt.close(fig)

    return classification, chart_path, rsi_val, macd_val, sig_val, detected_pattern, strength

# === Excel read/update ===
wb = openpyxl.load_workbook('stocks.xlsx')
ws_current = wb['Current']
if 'History' not in wb.sheetnames:
    ws_history = wb.create_sheet('History')
    ws_history.append(['Date','Ticker','Period','Freq','Classification','RSI','MACD','Signal','Pattern'])
else:
    ws_history = wb['History']
today = datetime.now().strftime('%Y-%m-%d')

for row in ws_current.iter_rows(min_row=2, max_col=3):
    tkr = row[0].value; per = row[1].value or '6mo'; freq = row[2].value or 'daily'
    if not tkr: continue
    try:
        cl, path, r, m, s, patt, strg = analyze_stock(tkr, per, freq)
        ws_current.cell(row=row[0].row, column=4).value = cl
        ws_current.cell(row=row[0].row, column=5).value = path
        ws_current.cell(row=row[0].row, column=6).value = patt
        ws_current.cell(row=row[0].row, column=7).value = strg
        if path and os.path.exists(path):
            img = XLImage(path); img.width=180; img.height=120
            ws_current.add_image(img, f'H{row[0].row}')
        ws_history.append([today,tkr,per,freq,cl,round(r,2),round(m,2),round(s,2),patt])
    except Exception as e:
        ws_current.cell(row=row[0].row, column=4).value = f'Error: {e}'
        print(f'⚠️ {tkr} error: {e}')
wb.save('stocks.xlsx')

# === PDF Report ===
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font('Arial','B',16)
pdf.cell(0,10,'Stock Pattern Summary Report',0,1,'C')
pdf.ln(5)
pdf.set_font('Arial',size=12)
pdf.cell(30,10,'Date:'); pdf.cell(50,10,today,0,1)
pdf.ln(5)

pdf.set_font('Arial','B',11)
pdf.cell(30,10,'Ticker',1)
pdf.cell(25,10,'Period',1)
pdf.cell(25,10,'Freq',1)
pdf.cell(30,10,'Classif',1)
pdf.cell(30,10,'Pattern',1)
pdf.ln()
pdf.set_font('Arial',size=10)
for vals in ws_current.iter_rows(min_row=2, max_col=7, values_only=True):
    tkr, period, freq, cl, chart_path, patt, strength = vals
    if tkr:
        pdf.cell(30,10,str(tkr),1)
        pdf.cell(25,10,str(period),1)
        pdf.cell(25,10,str(freq),1)
        pdf.cell(30,10,str(cl),1)
        pdf.cell(30,10,str(patt),1)
        pdf.ln()

# === Embed detailed charts, occurrences, and Top 20 ===
for idx, row in enumerate(ws_current.iter_rows(min_row=2, max_col=7), start=2):
    ticker = row[0].value
    chart_path = row[4].value
    pattern = row[5].value

    if not (ticker and pattern and chart_path and os.path.exists(chart_path)):
        continue

    pdf.add_page()
    pdf.set_font('Arial','B',14)
    pdf.cell(0,10,f"{ticker} | Pattern: {pattern}",0,1,'C')
    pdf.image(chart_path,10,30,190)

    period = ws_current.cell(row=idx, column=2).value or '6mo'
    freq = ws_current.cell(row=idx, column=3).value or 'daily'
    interval = {'daily':'1d', 'weekly':'1wk', 'monthly':'1mo'}.get(freq.lower(), '1d')

    df_full = yf.Ticker(ticker).history(period=period, interval=interval,
                                        auto_adjust=False, prepost=True)
    df_full = df_full[['Open','High','Low','Close']].dropna()
    df_full['Date_Num'] = mdates.date2num(df_full.index.to_pydatetime())

    df_occ = df_full.tail(250 if interval=='1d' else 52 if interval=='1wk' else 24) if str(period).lower() != 'max' else df_full

    talib_func = getattr(talib, 'CDL' + pattern, None)
    occ_idx = []
    if talib_func:
        result = talib_func(df_full['Open'], df_full['High'], df_full['Low'], df_full['Close'])
        occ_idx = np.where(result != 0)[0]

    occ_idx_plot = [i for i in occ_idx if i >= len(df_full) - len(df_occ)]

    fig2, ax2 = plt.subplots(figsize=(8,4))
    ax2.plot(df_occ.index, df_occ['Close'], label='Close')
    if occ_idx_plot:
        occ_dates = df_occ.index[[i - (len(df_full) - len(df_occ)) for i in occ_idx_plot]]
        occ_prices = df_occ['Close'].iloc[[i - (len(df_full) - len(df_occ)) for i in occ_idx_plot]]
        ax2.scatter(occ_dates, occ_prices, color=ACCENT, label=pattern)
    ax2.set_title(f"{ticker} Occurrences of {pattern}")
    ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    ax2.legend()
    fig2.autofmt_xdate()
    tmpfile = f"charts/{ticker}_occ_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png"
    fig2.savefig(tmpfile, dpi=100)
    plt.close(fig2)

    pdf.add_page()
    pdf.set_font('Arial','B',12)
    pdf.cell(0,10,f"{ticker} Pattern Occurrences: {len(occ_idx)}",0,1,'C')
    pdf.image(tmpfile,10,30,190)
    os.remove(tmpfile)

    pattern_counts = {}
    pattern_returns = {}
    for name in dir(talib):
        if name.startswith('CDL'):
            func = getattr(talib, name)
            result = func(df_full['Open'], df_full['High'], df_full['Low'], df_full['Close'])
            idxs = np.where(result != 0)[0]

            fwd_1y = []
            fwd_2y = []

            for idx_p in idxs:
                if idx_p + 252 < len(df_full):
                    start = df_full['Close'].iloc[idx_p]
                    end_1y = df_full['Close'].iloc[idx_p + 252]
                    fwd_1y.append((end_1y - start) / start)

                if idx_p + 504 < len(df_full):
                    start = df_full['Close'].iloc[idx_p]
                    end_2y = df_full['Close'].iloc[idx_p + 504]
                    fwd_2y.append((end_2y - start) / start)

            if len(idxs) > 0:
                pattern_name = name.replace('CDL','')
                pattern_counts[pattern_name] = int(len(idxs))
                pattern_returns[pattern_name] = {
                    '1Y': np.mean(fwd_1y)*100 if fwd_1y else None,
                    '2Y': np.mean(fwd_2y)*100 if fwd_2y else None
                }

    top_patterns = sorted(pattern_counts.items(), key=lambda x: x[1], reverse=True)[:20]

    pdf.add_page()
    pdf.set_font('Arial','B',12)
    pdf.cell(0,10,f"{ticker} | Top 20 Pattern Frequencies",0,1,'C')
    pdf.ln(4)

    pdf.set_font('Arial','B',10)
    pdf.cell(60,8,'Pattern Name',1)
    pdf.cell(25,8,'Occurrences',1)
    pdf.cell(30,8,'1Y Avg %',1)
    pdf.cell(30,8,'2Y Avg %',1)
    pdf.ln()

    pdf.set_font('Arial','',10)
    for pname, pcount in top_patterns:
        pdf.cell(60,8,str(pname),1)
        pdf.cell(25,8,str(pcount),1)
        r1 = pattern_returns.get(pname, {}).get('1Y')
        r2 = pattern_returns.get(pname, {}).get('2Y')
        pdf.cell(30,8, f"{r1:.2f}%" if r1 is not None else '-',1)
        pdf.cell(30,8, f"{r2:.2f}%" if r2 is not None else '-',1)
        pdf.ln()

pdf.output('stock_report.pdf')
print('✅ Analysis complete. PDF report saved as stock_report.pdf')

